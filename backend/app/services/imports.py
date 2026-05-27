import csv
import json
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.exceptions import AppError
from app.models.imports import ImportJob, ImportJobRow, ImportJobStatus, ImportRowStatus, ProductImportMode
from app.models.master_data import RecordStatus
from app.repositories.imports import ImportRepository

REQUIRED_FIELDS = {"name", "sku", "unit"}
OPTIONAL_FIELDS = {
    "barcode",
    "description",
    "category_name",
    "brand_name",
    "vendor_name",
    "cost_price",
    "selling_price",
    "reorder_level",
    "track_batch",
    "track_expiry",
    "track_serial",
    "status",
}
BOOLEAN_VALUES = {"true": True, "yes": True, "1": True, "false": False, "no": False, "0": False, "": False}
HEADER_ALIASES = {
    "product_name": "name",
    "item_name": "name",
    "item_code": "sku",
    "product_code": "sku",
    "hsn_code": "barcode",
    "category": "category_name",
    "brand": "brand_name",
    "vendor": "vendor_name",
    "cost": "cost_price",
    "price": "selling_price",
    "reorder": "reorder_level",
}


class ProductImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.repository = ImportRepository(db)

    @staticmethod
    def build_template_xlsx() -> bytes:
        headers = list(REQUIRED_FIELDS) + sorted(OPTIONAL_FIELDS)
        sample = [
            "Sample Product" if h == "name" else
            "SKU-001" if h == "sku" else
            "pcs" if h == "unit" else
            "" if h == "barcode" else
            "A sample product" if h == "description" else
            "Electronics" if h == "category_name" else
            "Acme Brand" if h == "brand_name" else
            "Sample Vendor" if h == "vendor_name" else
            "100.00" if h == "cost_price" else
            "150.00" if h == "selling_price" else
            "10" if h == "reorder_level" else
            "false" if h in ("track_batch", "track_expiry", "track_serial") else
            "active" if h == "status" else ""
            for h in headers
        ]
        try:
            import openpyxl
            import io
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products Import"
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for col, value in enumerate(sample, 1):
                ws.cell(row=2, column=col, value=value)
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
        except ImportError:
            pass

        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        ct_ns = "http://schemas.openxmlformats.org/package/2006/content-types"

        all_strings = headers + sample
        shared_strings_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        shared_strings_xml += f'<sst xmlns="{ns}" count="{len(all_strings)}" uniqueCount="{len(all_strings)}">'
        for s in all_strings:
            shared_strings_xml += f"<si><t>{s}</t></si>"
        shared_strings_xml += "</sst>"

        sheet_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        sheet_xml += f'<worksheet xmlns="{ns}"><sheetData><row r="1">'
        for idx, _ in enumerate(headers):
            col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
            sheet_xml += f'<c r="{col_letter}1" t="s"><v>{idx}</v></c>'
        sheet_xml += '</row><row r="2">'
        for idx, _ in enumerate(sample):
            col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
            sheet_xml += f'<c r="{col_letter}2" t="s"><v>{len(headers) + idx}</v></c>'
        sheet_xml += "</row></sheetData></worksheet>"

        workbook_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        workbook_xml += f'<workbook xmlns="{ns}" xmlns:r="{rel_ns}"><sheets><sheet name="Products Import" sheetId="1" r:id="rId1"/></sheets></workbook>'

        workbook_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        workbook_rels += '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        workbook_rels += '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        workbook_rels += '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        workbook_rels += "</Relationships>"

        rels_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        rels_xml += '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        rels_xml += '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        rels_xml += "</Relationships>"

        content_types = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        content_types += f'<Types xmlns="{ct_ns}">'
        content_types += '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        content_types += '<Default Extension="xml" ContentType="application/xml"/>'
        content_types += '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        content_types += '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        content_types += '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        content_types += "</Types>"

        buf = BytesIO()
        with ZipFile(buf, "w") as zf:
            zf.writestr("[Content_Types].xml", content_types)
            zf.writestr("_rels/.rels", rels_xml)
            zf.writestr("xl/workbook.xml", workbook_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
            zf.writestr("xl/sharedStrings.xml", shared_strings_xml)
        return buf.getvalue()

    def upload(
        self,
        tenant_id: int,
        actor_id: int,
        filename: str,
        content: bytes,
        mode: ProductImportMode,
        create_missing_references: bool,
        column_mapping_json: str | None = None,
    ) -> ImportJob:
        rows = self._parse_rows(filename, content, column_mapping_json)
        job = self.repository.create_job(
            {
                "tenant_id": tenant_id,
                "created_by": actor_id,
                "import_type": "products",
                "filename": filename,
                "mode": mode,
                "status": ImportJobStatus.UPLOADED,
                "total_rows": len(rows),
                "valid_rows": 0,
                "error_rows": 0,
                "warning_rows": 0,
                "created_count": 0,
                "updated_count": 0,
                "skipped_count": 0,
            }
        )
        self.repository.create_rows(
            [
                ImportJobRow(
                    tenant_id=tenant_id,
                    job_id=job.id,
                    row_number=index + 2,
                    raw_data={**row, "_create_missing_references": str(create_missing_references).lower()},
                    normalized_data=None,
                    status=ImportRowStatus.PENDING,
                    errors=[],
                    warnings=[],
                )
                for index, row in enumerate(rows)
            ]
        )
        self.db.commit()
        self.db.refresh(job)
        return job

    def get_job(self, tenant_id: int, job_id: int) -> ImportJob:
        job = self.repository.get_job(tenant_id, job_id)
        if job is None:
            raise AppError("IMPORT_JOB_NOT_FOUND", "Import job was not found for this tenant.", 404)
        return job

    def list_rows(self, tenant_id: int, job_id: int) -> list[ImportJobRow]:
        self.get_job(tenant_id, job_id)
        return self.repository.list_rows(tenant_id, job_id)

    def validate(self, tenant_id: int, job_id: int) -> tuple[ImportJob, list[ImportJobRow]]:
        job = self.get_job(tenant_id, job_id)
        if job.status in (ImportJobStatus.COMMITTED, ImportJobStatus.CANCELLED):
            raise AppError("INVALID_IMPORT_STATE", "Committed or cancelled jobs cannot be validated.", 409)
        job.status = ImportJobStatus.VALIDATING
        rows = self.repository.list_rows(tenant_id, job_id)
        seen_skus: dict[str, int] = {}
        seen_barcodes: dict[str, int] = {}
        create_missing = self._row_create_missing(rows)
        counts = {"valid": 0, "error": 0, "warning": 0}
        for row in rows:
            normalized, errors, warnings, existing_product_id = self._validate_row(tenant_id, row, job.mode, create_missing, seen_skus, seen_barcodes)
            row.normalized_data = normalized
            row.errors = errors
            row.warnings = warnings
            row.existing_product_id = existing_product_id
            if errors:
                row.status = ImportRowStatus.ERROR
                counts["error"] += 1
            elif warnings:
                row.status = ImportRowStatus.WARNING
                counts["warning"] += 1
                counts["valid"] += 1
            else:
                row.status = ImportRowStatus.VALID
                counts["valid"] += 1
        job.valid_rows = counts["valid"]
        job.error_rows = counts["error"]
        job.warning_rows = counts["warning"]
        job.status = ImportJobStatus.HAS_ERRORS if counts["error"] else ImportJobStatus.VALIDATED
        job.validated_at = datetime.now(UTC)
        self.db.commit()
        return job, rows

    def commit(self, tenant_id: int, job_id: int) -> tuple[ImportJob, list[ImportJobRow]]:
        job = self.get_job(tenant_id, job_id)
        if job.status == ImportJobStatus.UPLOADED:
            job, _ = self.validate(tenant_id, job_id)
        if job.status not in (ImportJobStatus.VALIDATED, ImportJobStatus.HAS_ERRORS):
            raise AppError("INVALID_IMPORT_STATE", "Only validated import jobs can be committed.", 409)
        created = updated = skipped = 0
        create_missing = self._row_create_missing(self.repository.list_rows(tenant_id, job_id))
        for row in self.repository.list_rows(tenant_id, job_id):
            if row.status == ImportRowStatus.ERROR or not row.normalized_data:
                row.status = ImportRowStatus.SKIPPED
                skipped += 1
                continue
            product = self.repository.get_product_by_sku(tenant_id, row.normalized_data["sku"])
            product_values = self._product_values(tenant_id, row.normalized_data, create_missing, row.warnings)
            if product:
                for key, value in product_values.items():
                    if key != "tenant_id":
                        setattr(product, key, value)
                row.status = ImportRowStatus.UPDATED
                row.existing_product_id = product.id
                updated += 1
            else:
                product = self.repository.create_product(product_values)
                row.status = ImportRowStatus.CREATED
                row.created_product_id = product.id
                created += 1
        job.created_count = created
        job.updated_count = updated
        job.skipped_count = skipped
        job.status = ImportJobStatus.COMMITTED
        job.committed_at = datetime.now(UTC)
        try:
            self.db.commit()
        except IntegrityError as exc:
            self.db.rollback()
            raise AppError("IMPORT_COMMIT_FAILED", "Import commit failed because of duplicate catalog data.", 409) from exc
        return job, self.repository.list_rows(tenant_id, job_id)

    def cancel(self, tenant_id: int, job_id: int) -> ImportJob:
        job = self.get_job(tenant_id, job_id)
        if job.status == ImportJobStatus.COMMITTED:
            raise AppError("INVALID_IMPORT_STATE", "Committed import jobs cannot be cancelled.", 409)
        job.status = ImportJobStatus.CANCELLED
        job.cancelled_at = datetime.now(UTC)
        self.db.commit()
        self.db.refresh(job)
        return job

    def _parse_rows(self, filename: str, content: bytes, column_mapping_json: str | None = None) -> list[dict[str, str]]:
        mapping = self._parse_mapping(column_mapping_json)
        lower_filename = filename.lower()
        if lower_filename.endswith(".xlsx") or content[:2] == b"PK":
            return self._parse_xlsx(content, mapping)
        return self._parse_csv(content, mapping)

    def _parse_mapping(self, column_mapping_json: str | None) -> dict[str, str]:
        if not column_mapping_json:
            return {}
        try:
            raw = json.loads(column_mapping_json)
        except json.JSONDecodeError as exc:
            raise AppError("INVALID_COLUMN_MAPPING", "Column mapping JSON is invalid.", 400) from exc
        if not isinstance(raw, dict):
            raise AppError("INVALID_COLUMN_MAPPING", "Column mapping JSON must be an object.", 400)
        return {str(source).strip().lower(): str(target).strip().lower() for source, target in raw.items() if source and target}

    def _canonical_header(self, value: str, mapping: dict[str, str]) -> str:
        normalized = value.strip().lower().replace(" ", "_").replace("-", "_")
        if normalized in mapping:
            return mapping[normalized]
        return HEADER_ALIASES.get(normalized, normalized)

    def _parse_csv(self, content: bytes, mapping: dict[str, str]) -> list[dict[str, str]]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise AppError("INVALID_IMPORT_FILE", "CSV file must be UTF-8 encoded.", 400) from exc
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise AppError("INVALID_IMPORT_FILE", "CSV file must include a header row.", 400)
        normalized_headers = [self._canonical_header(field or "", mapping) for field in reader.fieldnames]
        missing = REQUIRED_FIELDS - set(normalized_headers)
        if missing:
            raise AppError("INVALID_IMPORT_COLUMNS", f"Missing required columns: {', '.join(sorted(missing))}.", 400)
        rows = []
        for row in reader:
            rows.append({self._canonical_header((key or ""), mapping): (value or "").strip() for key, value in row.items()})
        return rows

    def _parse_xlsx(self, content: bytes, mapping: dict[str, str]) -> list[dict[str, str]]:
        try:
            with ZipFile(BytesIO(content)) as archive:
                shared_strings = self._shared_strings(archive)
                sheet_name = "xl/worksheets/sheet1.xml"
                if sheet_name not in archive.namelist():
                    raise AppError("INVALID_IMPORT_FILE", "XLSX workbook must include a first worksheet.", 400)
                root = ET.fromstring(archive.read(sheet_name))
        except AppError:
            raise
        except Exception as exc:
            raise AppError("INVALID_IMPORT_FILE", "Unable to parse the uploaded XLSX workbook.", 400) from exc
        namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rows = []
        header: list[str] | None = None
        for row in root.findall(".//a:sheetData/a:row", namespace):
            cells = [self._cell_value(cell, shared_strings, namespace).strip() for cell in row.findall("a:c", namespace)]
            if not any(cells):
                continue
            if header is None:
                header = [self._canonical_header(cell, mapping) for cell in cells]
                missing = REQUIRED_FIELDS - set(header)
                if missing:
                    raise AppError("INVALID_IMPORT_COLUMNS", f"Missing required columns: {', '.join(sorted(missing))}.", 400)
                continue
            record = {}
            for index, value in enumerate(cells):
                if index >= len(header):
                    continue
                record[header[index]] = value
            rows.append(record)
        return rows

    def _shared_strings(self, archive: ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in archive.namelist():
            return []
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        values = []
        for item in root.findall(".//a:si", namespace):
            text = "".join(node.text or "" for node in item.findall(".//a:t", namespace))
            values.append(text)
        return values

    def _cell_value(self, cell: ET.Element, shared_strings: list[str], namespace: dict[str, str]) -> str:
        cell_type = cell.attrib.get("t")
        value_node = cell.find("a:v", namespace)
        inline_text = cell.find("a:is/a:t", namespace)
        if inline_text is not None:
            return inline_text.text or ""
        if value_node is None or value_node.text is None:
            return ""
        raw = value_node.text
        if cell_type == "s":
            try:
                return shared_strings[int(raw)]
            except (IndexError, ValueError):
                return ""
        return raw

    def _validate_row(self, tenant_id: int, row: ImportJobRow, mode: ProductImportMode, create_missing: bool, seen_skus: dict[str, int], seen_barcodes: dict[str, int]) -> tuple[dict[str, Any], list[str], list[str], int | None]:
        raw = row.raw_data
        errors: list[str] = []
        warnings: list[str] = []
        normalized: dict[str, Any] = {}
        for field in REQUIRED_FIELDS:
            value = raw.get(field, "").strip()
            if not value:
                errors.append(f"{field} is required")
            normalized[field] = value
        normalized["barcode"] = raw.get("barcode") or None
        normalized["description"] = raw.get("description") or None
        normalized["category_name"] = raw.get("category_name") or None
        normalized["brand_name"] = raw.get("brand_name") or None
        normalized["vendor_name"] = raw.get("vendor_name") or None
        normalized["status"] = raw.get("status") or RecordStatus.ACTIVE.value
        if normalized["status"] not in {item.value for item in RecordStatus}:
            errors.append("status must be ACTIVE, INACTIVE, or ARCHIVED")
        for price_field in ["cost_price", "selling_price"]:
            parsed = self._decimal(raw.get(price_field), price_field, errors)
            normalized[price_field] = str(parsed) if parsed is not None else None
        normalized["reorder_level"] = self._non_negative_int(raw.get("reorder_level"), "reorder_level", errors)
        for flag in ["track_batch", "track_expiry", "track_serial"]:
            normalized[flag] = self._bool(raw.get(flag, "false"), flag, errors)
        sku = normalized.get("sku")
        barcode = normalized.get("barcode")
        if sku:
            if sku in seen_skus:
                errors.append(f"duplicate SKU in file; first seen on row {seen_skus[sku]}")
            else:
                seen_skus[sku] = row.row_number
        if barcode:
            if barcode in seen_barcodes:
                errors.append(f"duplicate barcode in file; first seen on row {seen_barcodes[barcode]}")
            else:
                seen_barcodes[barcode] = row.row_number
        existing = self.repository.get_product_by_sku(tenant_id, sku) if sku else None
        if mode == ProductImportMode.create_only and existing:
            errors.append("SKU already exists for this tenant")
        if mode == ProductImportMode.update_existing and not existing:
            errors.append("SKU does not exist for update_existing mode")
        if barcode:
            barcode_product = self.repository.get_product_by_barcode(tenant_id, barcode)
            if barcode_product and (not existing or barcode_product.id != existing.id):
                errors.append("barcode already exists for this tenant")
        for ref_field in ["category_name", "brand_name", "vendor_name"]:
            name = normalized.get(ref_field)
            if name and not create_missing:
                warnings.append(f"{ref_field} will only be linked if it already exists")
        if normalized.get("vendor_name"):
            warnings.append("Vendor exists/created but product-vendor linking is not supported in this phase")
        return normalized, errors, warnings, existing.id if existing else None

    def _product_values(self, tenant_id: int, data: dict[str, Any], create_missing: bool, warnings: list[str]) -> dict[str, Any]:
        values = {
            "tenant_id": tenant_id,
            "name": data["name"],
            "sku": data["sku"],
            "barcode": data.get("barcode"),
            "description": data.get("description"),
            "unit": data["unit"],
            "cost_price": Decimal(str(data["cost_price"])) if data.get("cost_price") is not None else None,
            "selling_price": Decimal(str(data["selling_price"])) if data.get("selling_price") is not None else None,
            "reorder_level": data.get("reorder_level"),
            "track_batch": data.get("track_batch", False),
            "track_expiry": data.get("track_expiry", False),
            "track_serial": data.get("track_serial", False),
            "status": RecordStatus(data.get("status", RecordStatus.ACTIVE.value)),
        }
        if data.get("category_name"):
            category = self.repository.get_category_by_name(tenant_id, data["category_name"])
            if category is None and create_missing:
                category = self.repository.create_category(tenant_id, data["category_name"])
            values["category_id"] = category.id if category else None
        if data.get("brand_name"):
            brand = self.repository.get_brand_by_name(tenant_id, data["brand_name"])
            if brand is None and create_missing:
                brand = self.repository.create_brand(tenant_id, data["brand_name"])
            values["brand_id"] = brand.id if brand else None
        if data.get("vendor_name") and create_missing and self.repository.get_vendor_by_name(tenant_id, data["vendor_name"]) is None:
            self.repository.create_vendor(tenant_id, data["vendor_name"])
        return values

    def _row_create_missing(self, rows: list[ImportJobRow]) -> bool:
        return any(row.raw_data.get("_create_missing_references") == "true" for row in rows)

    def _decimal(self, value: str | None, field: str, errors: list[str]) -> Decimal | None:
        if not value:
            return None
        try:
            parsed = Decimal(str(value))
        except (InvalidOperation, ValueError):
            errors.append(f"{field} must be numeric")
            return None
        if parsed < 0:
            errors.append(f"{field} cannot be negative")
        return parsed

    def _non_negative_int(self, value: str | None, field: str, errors: list[str]) -> int | None:
        if not value:
            return None
        try:
            parsed = int(value)
        except ValueError:
            errors.append(f"{field} must be an integer")
            return None
        if parsed < 0:
            errors.append(f"{field} cannot be negative")
        return parsed

    def _bool(self, value: str | None, field: str, errors: list[str]) -> bool:
        key = (value or "").strip().lower()
        if key not in BOOLEAN_VALUES:
            errors.append(f"{field} must be boolean-compatible")
            return False
        return BOOLEAN_VALUES[key]
